"""
This is the script which is trying to gather all the results in one location from all the 
results file. 

Author : Nikhil Kondabala
Date: 01/16/2018 14:51
Version: 0.0.1 

the general process that we are trying to get to is get the filename of the resutls , gather company 
project details. then gather all the regression resutls in one locaiton 
get all the other attributes in one location and make a dataframe out of it. 
then read the excel file, get the last row of the excel sheet and then enter the data into the excel sheet. 

usage DataParsingResults.py 


"""

#TODO: Add headers to the output file. 
#TODO: Think about the clean up for the empty rows 
#TODO: get the count details in one location
#TODO: currently this is only setup to run in Nikhil's computer need to change that

import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

def write_resultstofile(df,ws):
    # write the regression results to file.
    rows = dataframe_to_rows(df, index=True, header=True)
    for r in rows:
        ws.append(r)

def get_datasplitout(filename):
    df = pd.read_excel(filename)
    # getting the correct column keys 
    cols_key_df = pd.read_excel('./Column_Key.xlsx')
    df_cols = df.columns.to_frame()
    df_cols.index = range(0,78)
    df_cols = df_cols.join(cols_key_df)
    df_cols_dicts = df_cols.to_dict('records')

    # removing the null values. 
    data=df[~df.iloc[:,0].isnull() | ~df.index.isnull()]
    f = filename.split('\\')[-1]

    company= f.split('_')[2]
    project= f.split('_')[3]
    attribute = data.index.unique()
    attribute = attribute[~attribute.isnull()]
    
    def add_projectdetails(df):
        df = df.assign(Project=project)
        df= df.assign(Company=company)
        return df

    def order_cols_df(df,Col_touse):
        cols = list(df.columns)
        i = [len(cols)-1, len(cols)-2]
        i+=range(0,len(cols)-2)
        orderedcols = [cols[k] for k in i]
        df = df[orderedcols]# this is the new dataframe with ordered cols
        
        newcols = []
        for col in orderedcols:
            lookup = next((item for item in df_cols_dicts if item["Original"] == col), None)
            if not lookup:
                lookup = {Col_touse:col}
            newcols.append(lookup[Col_touse])
        df.columns = newcols # this is the new dataframe with renamed cols

        return df

    # TRY TO GET ALL THE RESULTS WE NEED IN A DATAFRAME 
    def get_aboveandbelow_nominal_stats():
        ab_row = data.index.get_loc('Above nominal Status')
        b_row = data.index.get_loc('Below nominal Stats')
        a_row = data.index.get_loc('Total Bin Stats')
        cols = ['TI_error_mean',	'TI_error_std',	'TI_diff_mean',	'TI_diff_std']

        ab_stats = data.iloc[ab_row+2:ab_row+6,0:4]
        b_stats = data.iloc[b_row+2:b_row+6,0:4]
        a_stats = data.iloc[a_row+2:a_row+6,0:4]
        
        ab_stats.columns = [i+'_AboveNominal' for i in cols]
        b_stats.columns = [i+'_BelowNominal' for i in cols]
        a_stats.columns = [i+'_All' for i in cols]

        stats = ab_stats.join(b_stats)
        stats = stats.join(a_stats)
        TI_diff_cols = stats.columns[stats.columns.str.contains('TI_diff')].tolist()
        TI_error_cols = stats.columns[stats.columns.str.contains('TI_error')].tolist()
        TI_diff_df , TI_error_df = stats.loc[:,TI_diff_cols] , stats.loc[:,TI_error_cols]

        diff_rows = TI_diff_df.index.tolist()
        diff_rows = ['TI_diff_'+i for i in diff_rows]
        TI_diff_df.index = diff_rows

        error_rows = TI_error_df.index.tolist()
        error_rows = ['TI_error_'+i for i in error_rows]
        TI_error_df.index = error_rows

        return TI_diff_df , TI_error_df    
    
    statistics_diff, statistics_error = get_aboveandbelow_nominal_stats()

    def get_regressionresults(): #REGRESSION RESULTS IN A DATAFRAME 
        reggression = attribute[attribute.str.contains('WS_regression')]
        reggression_df = data.loc[reggression].copy()
        reggression_df = reggression_df.iloc[:,0:4]

        reggression_df = add_projectdetails(reggression_df)
        reggression_df = order_cols_df(reggression_df, 'Original')

        return reggression_df

    def get_TI_error_results(): # get the TI mean bias error results by bin    
        TI_error =attribute[attribute.str.contains('TI_error')]
        TI_error_df = data.loc[TI_error].copy()

        rows = TI_error_df.shape[0]
        TI_error_df_1mps = TI_error_df.iloc[range(0,rows,2),:]
        TI_error_df_1mps = TI_error_df_1mps.join(statistics_error)
        TI_error_df_05mps = TI_error_df.iloc[range(1,rows,2),:]
        TI_error_df_05mps = TI_error_df_05mps.join(statistics_error)
        # add project details and order cols 
        TI_error_df_1mps = add_projectdetails(TI_error_df_1mps)
        TI_error_df_1mps = order_cols_df(TI_error_df_1mps, '1mps_bins')
        TI_error_df_05mps = add_projectdetails(TI_error_df_05mps)
        TI_error_df_05mps = order_cols_df(TI_error_df_05mps, 'p5mps_bins')

       
        return TI_error_df_1mps , TI_error_df_05mps

    def get_TI_count_results(): # get the TI mean bias error results by bin    
        TI_count =attribute[attribute.str.contains('RSD_WS')]
        TI_count_df = data.loc[TI_count].copy()

        rows = TI_count_df.shape[0]
        TI_count_df_1mps = TI_count_df.iloc[range(0,rows,2),:]
        TI_count_df_1mps = TI_count_df_1mps.join(statistics_error)
        TI_count_df_05mps = TI_count_df.iloc[range(1,rows,2),:]
        TI_count_df_05mps = TI_count_df_05mps.join(statistics_error)
        # add project details and order cols 
        TI_count_df_1mps = add_projectdetails(TI_count_df_1mps)
        TI_count_df_1mps = order_cols_df(TI_count_df_1mps, '1mps_bins')
        TI_count_df_05mps = add_projectdetails(TI_count_df_05mps)
        TI_count_df_05mps = order_cols_df(TI_count_df_05mps, 'p5mps_bins')

       
        return TI_count_df_1mps , TI_count_df_05mps

    def get_TI_diff_results(): # get the TI_difference results based on bin 
        TI_diff =attribute[attribute.str.contains('TI_diff')]
        TI_diff_df = data.loc[TI_diff].copy()

        rows = TI_diff_df.shape[0]
        TI_diff_df_1mps = TI_diff_df.iloc[range(0,rows,2),:]
        TI_diff_df_1mps = TI_diff_df_1mps.join(statistics_diff)
        TI_diff_df_05mps = TI_diff_df.iloc[range(1,rows,2),:]
        TI_diff_df_05mps = TI_diff_df_05mps.join(statistics_diff)

        TI_diff_df_05mps = add_projectdetails(TI_diff_df_05mps)
        TI_diff_df_1mps = add_projectdetails(TI_diff_df_1mps)

        TI_diff_df_1mps = order_cols_df(TI_diff_df_1mps, '1mps_bins')
        TI_diff_df_05mps = order_cols_df(TI_diff_df_05mps, 'p5mps_bins')

        return TI_diff_df_1mps , TI_diff_df_05mps 

    def get_TI_values():# get the TI values by bin, get the representative TI values based on mean TI and Std
        TI_values = attribute[attribute.str.contains('_TI')]
        TI_values_df=data.loc[TI_values].copy()
        TI_values_df = add_projectdetails(TI_values_df)

        totallen = TI_values_df.shape[0]
        numofvar = TI_values.shape[0]

        # make a dataframe with aggregate values for the TI values for representative TI
        TI_values_agg_df = TI_values_df.iloc[range(2,totallen,numofvar),0:3]
        TI_values_agg_df.columns = ['mean_15mps','std_15mps','Rep_TI']
        # get the 1 mps bin TI values and add the aggregate representative ti values to the end of df
        TI_values_1mpsbin_df = TI_values_df.iloc[range(0,totallen,numofvar),:]
        # get the .05 mps bin TI values and add the aggregate representative ti values to the end of df
        TI_values_05mpsbin_df = TI_values_df.iloc[range(1,totallen,numofvar),:]
        TI_values_05mpsbin_df = order_cols_df(TI_values_05mpsbin_df, 'p5mps_bins')
        TI_values_1mpsbin_df = order_cols_df(TI_values_1mpsbin_df, '1mps_bins')
        TI_values_agg_df = add_projectdetails(TI_values_agg_df)
        TI_values_agg_df = order_cols_df(TI_values_agg_df, 'Original')

        return TI_values_agg_df, TI_values_1mpsbin_df, TI_values_05mpsbin_df
    
    regression_stats = get_regressionresults()
    TI_values_agg_df, TI_values_1mpsbin_df, TI_values_05mpsbin_df = get_TI_values()
    TI_diff_df_1mps , TI_diff_df_05mps = get_TI_diff_results()
    TI_error_df_1mps , TI_error_df_05mps = get_TI_error_results()
    TI_count_df_1mps , TI_count_df_05mps =get_TI_count_results()

    results_1mps = [regression_stats,TI_values_agg_df,TI_values_1mpsbin_df,TI_diff_df_1mps,TI_error_df_1mps, TI_count_df_1mps]
    results_05mps = [regression_stats,TI_values_agg_df,TI_values_05mpsbin_df,TI_diff_df_05mps,TI_error_df_05mps, TI_count_df_05mps]
    return results_1mps, results_05mps

def create_master_workbook():
    wb = Workbook()
    wb.create_sheet(title='Regression')
    wb.create_sheet(title='TI_MBE')
    wb.create_sheet(title='TI_diff')
    wb.create_sheet(title='TI_values')
    wb.create_sheet(title='TI_agg')
    wb.create_sheet(title='TI_count')
    return wb
def write_results_aggregate_workbook(wb, results, savefileas):
    # write the results to the file
    reg_sheet = wb['Regression']
    TI_MBE_sheet = wb['TI_MBE']
    TI_diff_sheet = wb['TI_diff']
    TI_values_sheet = wb['TI_values']
    TI_agg_sheet = wb['TI_agg']
    TI_count_sheet = wb['TI_count']
  
    write_resultstofile(results[0],reg_sheet)
    write_resultstofile(results[2],TI_values_sheet)
    write_resultstofile(results[3],TI_diff_sheet)
    write_resultstofile(results[1],TI_agg_sheet)
    write_resultstofile(results[4],TI_MBE_sheet)
    write_resultstofile(results[5],TI_count_sheet)
  

    wb.save(savefileas)


def get_listofFilestoread():
    import os
    cwd = os.getcwd()
    print cwd
    path = cwd+'\\results\\'
    files = []
    # r=root, d=directories, f = files
    for r, d, f in os.walk(path):
        for file in f:
            if '.xlsx' in file:
                files.append(os.path.join(r, file))

    return files, path

def cleanup_results(filename, wb, savefileas):
    #make a new worksheet work book object so that we can clean it up. 
    sheets = ['Regression','TI_MBE','TI_diff','TI_values','TI_agg','TI_count']
    for sheet in sheets:
        df = pd.read_excel(filename, sheet_name=sheet)
        df = df[~df.index.isna()]
        ws = wb[sheet]
        write_resultstofile(df,ws)
    wb.save(savefileas)

if __name__=="__main__":

    files, path = get_listofFilestoread()
    wb = create_master_workbook()
    wb1 = create_master_workbook()
    wb2 = create_master_workbook()
    wb3 = create_master_workbook()
    
    results_file_1mps = path+'CFARS_Aggregate_Results_Phase1test_1mps.xlsx'
    results_file_05mps = path+'CFARS_Aggregate_Results_Phase1test_05mps.xlsx'
    final_results_file_1mps = path+'Final_CFARS_Aggregate_Results_Phase1test_1mps.xlsx'
    final_results_file_05mps = path+'Final_CFARS_Aggregate_Results_Phase1test_05mps.xlsx'
    
    for f in files:
        try:
        # f = "C:/Users/nikhil.kondabala/Documents/GitHub/CFARS_SS/results/APEX/Phase1Tests_ResultsMatrix_Apex_10317_v01_20181130.xlsx"
            results_1mps, results_05mps = get_datasplitout(f)
            write_results_aggregate_workbook(wb,results_1mps,results_file_1mps)
            write_results_aggregate_workbook(wb1,results_05mps,results_file_05mps)
        except:
            print 'there is a error in the file:  {} '.format(f)
    
    cleanup_results(results_file_05mps, wb3, final_results_file_05mps)