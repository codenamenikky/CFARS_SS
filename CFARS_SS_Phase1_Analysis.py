"""
This is the main script which is trying to clean and input data in the correct order into the internal input sheet. 
this will be used for the Phase 1 test summrization. 
Author : Nikhil Kondabala
Date: 11/16/2018 14:19
Version: 0.0.1 
"""


import pandas as pd
# import statsmodels.formula.api as sm
# import statsmodels.api as sm_api 
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import argparse
from sklearn import linear_model 
def set_inputdataformat(config_file):
# this function uses the input data function given in a configration file and converts into a dict 
# of renamed cols which will be used to change the format from your internal format to a CFARS specific one 
#  so that the rest of the script can work 
    df = pd.read_excel(config_file)
    df = df.dropna()
    return dict(zip(df.Internal_Column.tolist(), df.CFARS_Column.tolist()))
def get_CFARScolumns(config_file):
    df = pd.read_excel(config_file)
    df = df.dropna()
    return df.CFARS_Column.tolist()

def get_inputdata(filename, config_file):
# This is the function to get the input data to analyze
# the input data can be internal structure and does not need to be a set format. 
#  the formatting information of the data is provided in the config file, the config file template is in the git hub repo

    if filename.split('.')[-1] == 'csv':
        inputdata = pd.read_csv(filename)  
    elif filename.split('.')[-1] == 'xlsx':
        inputdata = pd.read_excel(filename)
    else:
        print 'Unkown input file type for the input data , please consider changing it to csv'
        sys.exit()

    try:
        rename_cols = set_inputdataformat(config_file)
    except Exception as e:
        print 'There is an error in the configuration file'
        print e
        sys.exit()

    inputdata = inputdata.rename(index=str,columns=rename_cols)
    inputdata = inputdata.dropna() # this needs to be thought through, in essence we are applying a wholesale drop for non exisiting data 
    inputdata['bins']=inputdata['Ref_WS'].round(0) # this acts as bin because the bin defination is between the two half integer values 

    bins_p5_interval= pd.interval_range(start=.25,end=20,freq=.5, closed='left')# this is creating a interval range of .5 starting at .25
    out =pd.cut(x= inputdata['Ref_WS'], bins=bins_p5_interval)
    inputdata['bins_p5']=out.apply(lambda x: x.mid) # the middle of the interval is used as a catagorical label 
    inputdata = inputdata[inputdata['Ref_TI']!=0] # we can only analyze where the ref_TI is not 0

    return inputdata

#TODO: export the representative TI plot
# def plot_rep_TI(representative_TI, projectname):
#     # repcols = [u'RSD_TI_rep',u'Ref_TI_rep', u'Ane2_TI_rep']
#     # plt = representative_TI[repcols].plot(title='Rep TI plot, TI_mean_bin+TI_sd_bin*1.28', ylable='Rep TI')
#     # plt.figure.savefig('./{}_rep_TIplot.png'.format(projectname))

def get_regression_sm(x,y):
# get the linear least squaes fit, this uses statsmodel which needs to be installed outside of Anaconda 
    
    x = sm_api.add_constant(x)
    model = sm_api.OLS(y,x,missing='drop').fit()
    result = model.params
    result['Rsquared'] = model.rsquared
    result['WSdiff'] = abs((x.iloc[:,1]-y).mean())
    return [result[1], result[0], result[2], result[3]]

def get_regression(x,y):
# get the linear least squaes fit, this uses the sklearn model which is in Anaconda already.  
    lm = linear_model.LinearRegression()
    lm.fit(x.to_frame(),y.to_frame())
    result = [lm.coef_[0][0],lm.intercept_[0]]
    result.append(lm.score(x.to_frame(),y.to_frame()))
    result.append(abs((x-y).mean()))
    return result

def get_ws_regression(inputdata):
    # get the ws regression results for all the col required pairs. 
    results = pd.DataFrame(columns=['m','c','rsquared','ws_diff'])
    results_rsd_ref = get_regression(inputdata['RSD_WS'],inputdata['Ref_WS']) # this is to get the rsd and ref anemometer results. 
    results.loc['WS_regression_RSD_Ref'] = results_rsd_ref
    
    if 'corrTI_RSD_WS' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results. 
        results_corrTI_rsd_ref = get_regression(inputdata['corrTI_RSD_WS'],inputdata['Ref_WS'])
        results.loc['WS_regression_corrTI_RSD_Ref']=results_corrTI_rsd_ref
    else:
        results.loc['WS_regression_corrTI_RSD_Ref']=[None, None, None, None]
    if 'corrWS_RSD_WS' in inputdata.columns:
        results_corrWS_rsd_ref = get_regression(inputdata['corrTI_RSD_WS'],inputdata['Ref_WS'])
        results.loc['WS_regression_corrWS_RSD_Ref']=results_corrWS_rsd_ref
    else:
        results.loc['WS_regression_corrWS_RSD_Ref']=[None, None, None, None]
    
    results_Ane2_ref = get_regression(inputdata['Ane2_WS'],inputdata['Ref_WS'])
    results.loc['WS_regression_Ane2_Ref']=results_Ane2_ref        

    return results


def get_representative_TI(inputdata):
# get the representaive TI
    representative_TI = inputdata[['RSD_TI','Ref_TI','Ane2_TI','bins']].groupby(by=['bins']).agg(['mean','std',lambda x: x.mean()+1.28*x.std()])
    representative_TI.columns = ['RSD_TI_mean','RSD_TI_std','RSD_TI_rep','Ref_TI_mean','Ref_TI_std','Ref_TI_rep','Ane2_TI_mean','Ane2_TI_std','Ane2_TI_rep']
    return representative_TI

def get_count_per_WSbin(inputdata, column):
    # Count per wind speed bin
    inputdata = inputdata[(inputdata['bins_p5'].astype(float)>1.5) & (inputdata['bins_p5'].astype(float)<21)]  
    resultsstats_bin = inputdata[[column,'bins']].groupby(by = 'bins').agg(['count'])
    resultsstats_bin_p5 = inputdata[[column,'bins_p5']].groupby(by = 'bins_p5').agg(['count'])
    resultsstats_bin = pd.DataFrame(resultsstats_bin.unstack()).T
    resultsstats_bin.index = [column]
    resultsstats_bin_p5 = pd.DataFrame(resultsstats_bin_p5.unstack()).T
    resultsstats_bin_p5.index = [column]   
    return resultsstats_bin, resultsstats_bin_p5

def get_stats_per_WSbin(inputdata, column):
    # this will be used as a base function for all frequency agg caliculaitons for each bin to get the stats per wind speed bins
    inputdata = inputdata[(inputdata['bins_p5'].astype(float)>1.5) & (inputdata['bins_p5'].astype(float)<21)]  
    resultsstats_bin = inputdata[[column,'bins']].groupby(by = 'bins').agg(['mean','std'])
    resultsstats_bin_p5 = inputdata[[column,'bins_p5']].groupby(by = 'bins_p5').agg(['mean','std'])
    resultsstats_bin = pd.DataFrame(resultsstats_bin.unstack()).T
    resultsstats_bin.index = [column]
    resultsstats_bin_p5 = pd.DataFrame(resultsstats_bin_p5.unstack()).T
    resultsstats_bin_p5.index = [column]
    return resultsstats_bin, resultsstats_bin_p5 

def get_TI_MBE_Diff_j(inputdata):
    # get the bin wise stats for both diff and error between RSD and Ref
    inputdata['TI_diff_RSD_Ref']= inputdata['RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
    inputdata['TI_error_RSD_Ref']= inputdata['TI_diff_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
    TI_MBE_j_RSD_Ref,TI_MBE_jp5_RSD_Ref  = get_stats_per_WSbin(inputdata,'TI_error_RSD_Ref')
    TI_Diff_j_RSD_Ref, TI_Diff_jp5_RSD_Ref = get_stats_per_WSbin(inputdata,'TI_diff_RSD_Ref')
    TI_MBE_j_ = []
    TI_Diff_j_ = []

    TI_MBE_j_.append([TI_MBE_j_RSD_Ref,TI_MBE_jp5_RSD_Ref])
    TI_Diff_j_.append([TI_Diff_j_RSD_Ref, TI_Diff_jp5_RSD_Ref])
    # get the bin wise stats for both diff and error between RSD corrected for TI and Ref
    
    if 'corrTI_RSD_TI' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results. 
        inputdata['TI_diff_corrTI_RSD_Ref']= inputdata['corrTI_RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrTI_RSD_Ref']= inputdata['TI_diff_corrTI_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_MBE_j_corrTI_rsd_ref, TI_MBE_jp5_corrTI_rsd_ref = get_stats_per_WSbin(inputdata,'TI_error_corrTI_RSD_Ref')
        TI_Diff_j_corrTI_rsd_ref, TI_Diff_jp5_corrTI_rsd_ref = get_stats_per_WSbin(inputdata,'TI_diff_corrTI_RSD_Ref')
        TI_MBE_j_.append([TI_MBE_j_corrTI_rsd_ref, TI_MBE_jp5_corrTI_rsd_ref])
        TI_Diff_j_.append([ TI_Diff_j_corrTI_rsd_ref, TI_Diff_jp5_corrTI_rsd_ref])
    else:
        TI_MBE_j_.append(pd.DataFrame([None, None]))
        TI_Diff_j_.append(pd.DataFrame([None, None]))
    
    # get the bin wise stats for both diff and error between RSD corrected for ws and Ref
        
    if 'corrWS_RSD_TI' in inputdata.columns: #this is checking if the corrected WS method was used for the wind speed and TI
        inputdata['TI_diff_corrWS_RSD_Ref']= inputdata['corrWS_RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrWS_RSD_Ref']= inputdata['TI_diff_corrWS_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_MBE_j_corrWS_rsd_ref, TI_MBE_jp5_corrWS_rsd_ref = get_stats_per_WSbin(inputdata,'TI_error_corrWS_RSD_Ref')
        TI_Diff_j_corrWS_rsd_ref, TI_Diff_jp5_corrWS_rsd_ref = get_stats_per_WSbin(inputdata,'TI_diff_corrWS_RSD_Ref')
        TI_MBE_j_.append([TI_MBE_j_corrWS_rsd_ref, TI_MBE_jp5_corrWS_rsd_ref])
        TI_Diff_j_.append([ TI_Diff_j_corrWS_rsd_ref, TI_Diff_jp5_corrWS_rsd_ref])
    else:
        TI_MBE_j_.append(pd.DataFrame([None, None]))
        TI_Diff_j_.append(pd.DataFrame([None, None]))

    #get bin wise stats for the ane2 and ref anemometer 

    inputdata['TI_diff_Ane2_Ref']= inputdata['Ane2_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
    inputdata['TI_error_Ane2_Ref']= inputdata['TI_diff_Ane2_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
    TI_MBE_j_Ane2_Ref,TI_MBE_jp5_Ane2_Ref  = get_stats_per_WSbin(inputdata,'TI_error_Ane2_Ref')
    TI_Diff_j_Ane2_Ref, TI_Diff_jp5_Ane2_Ref = get_stats_per_WSbin(inputdata,'TI_diff_Ane2_Ref')
    TI_MBE_j_.append([TI_MBE_j_Ane2_Ref,TI_MBE_jp5_Ane2_Ref])
    TI_Diff_j_.append([ TI_Diff_j_Ane2_Ref, TI_Diff_jp5_Ane2_Ref])
    
    return TI_MBE_j_,TI_Diff_j_ 

def get_TI_bybin(inputdata):
    results = []
    RSD_TI_j, RSD_TI_jp5 = get_stats_per_WSbin(inputdata,'RSD_TI')
    results.append([RSD_TI_j, RSD_TI_jp5])
    
    Ref_TI_j, Ref_TI_jp5 = get_stats_per_WSbin(inputdata,'Ref_TI')
    results.append([Ref_TI_j, Ref_TI_jp5])

    if 'corrTI_RSD_TI' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results. 
        corrTI_RSD_TI_j, corrTI_RSD_TI_jp5 = get_stats_per_WSbin(inputdata,'corrTI_RSD_TI')
        results.append([corrTI_RSD_TI_j,corrTI_RSD_TI_jp5])
    else:
        results.append(pd.DataFrame([None, None]))

    # get the bin wise stats for both diff and error between RSD corrected for ws and Ref
        
    if 'corrWS_RSD_TI' in inputdata.columns: #this is checking if the corrected WS method was used for the wind speed and TI
        corrWS_RSD_TI_j, corrWS_RSD_TI_jp5 = get_stats_per_WSbin(inputdata,'corrWS_RSD_TI')
        results.append([corrWS_RSD_TI_j, corrWS_RSD_TI_jp5])
    else:
        results.append(pd.DataFrame([None, None]))

    Ane2_TI_j, Ane2_TI_jp5 = get_stats_per_WSbin(inputdata,'Ane2_TI')
    results.append([Ane2_TI_j, Ane2_TI_jp5])
    
    return results

def get_stats_inBin(inputdata_m,start,end):    
    # this was discussed in the meeting , but the results template didn't ask for this. 
    inputdata = inputdata_m.loc[(inputdata_m['Ref_WS']>start) & (inputdata_m['Ref_WS']<=end)].copy()
    inputdata['TI_diff_RSD_Ref']= inputdata['RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
    inputdata['TI_error_RSD_Ref']= inputdata['TI_diff_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
    TI_error_RSD_Ref_Avg = inputdata['TI_error_RSD_Ref'].mean()
    TI_error_RSD_Ref_Std = inputdata['TI_error_RSD_Ref'].std()
    TI_diff_RSD_Ref_Avg = inputdata['TI_diff_RSD_Ref'].mean()
    TI_diff_RSD_Ref_Std = inputdata['TI_diff_RSD_Ref'].std()
    results = pd.DataFrame([TI_error_RSD_Ref_Avg,TI_error_RSD_Ref_Std,TI_diff_RSD_Ref_Avg,TI_diff_RSD_Ref_Std],columns=['RSD_Ref'])
    if 'corrTI_RSD_TI' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results. 
        inputdata['TI_diff_corrTI_RSD_Ref']= inputdata['corrTI_RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrTI_RSD_Ref']= inputdata['TI_diff_corrTI_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_error_corrTI_RSD_Ref_Avg = inputdata['TI_error_corrTI_RSD_Ref'].mean()
        TI_error_corrTI_RSD_Ref_Std = inputdata['TI_error_corrTI_RSD_Ref'].std()
        TI_diff_corrTI_RSD_Ref_Avg = inputdata['TI_diff_corrTI_RSD_Ref'].mean()
        TI_diff_corrTI_RSD_Ref_Std = inputdata['TI_diff_corrTI_RSD_Ref'].std()
        results['CorrTI_RSD_Ref'] = [TI_error_corrTI_RSD_Ref_Avg,TI_error_corrTI_RSD_Ref_Std, TI_diff_corrTI_RSD_Ref_Avg, TI_diff_corrTI_RSD_Ref_Std]
    else:
        results['CorrTI_RSD_Ref'] = [None, None, None, None]
    # get the bin wise stats for both diff and error between RSD corrected for ws and Ref
        
    if 'corrWS_RSD_TI' in inputdata.columns: #this is checking if the corrected WS method was used for the wind speed and TI
        inputdata['TI_diff_corrWS_RSD_Ref']= inputdata['corrWS_RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrWS_RSD_Ref']= inputdata['TI_diff_corrWS_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_error_corrWS_RSD_Ref_Avg = inputdata['TI_error_corrWS_RSD_Ref'].mean()
        TI_error_corrWS_RSD_Ref_Std = inputdata['TI_error_corrWS_RSD_Ref'].std()
        TI_diff_corrWS_RSD_Ref_Avg = inputdata['TI_diff_corrWS_RSD_Ref'].mean()
        TI_diff_corrWS_RSD_Ref_Std = inputdata['TI_diff_corrWS_RSD_Ref'].std()
        results['corrWS_RSD_Ref'] = [TI_error_corrWS_RSD_Ref_Avg,TI_error_corrWS_RSD_Ref_Std,TI_diff_corrWS_RSD_Ref_Avg,TI_diff_corrWS_RSD_Ref_Std]
    else:
        results['corrWS_RSD_Ref'] =[None, None, None, None]
    
    inputdata['TI_diff_Ane2_Ref']= inputdata['Ane2_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
    inputdata['TI_error_Ane2_Ref']= inputdata['TI_diff_Ane2_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
    TI_error_Ane2_Ref_Avg = inputdata['TI_error_Ane2_Ref'].mean()
    TI_error_Ane2_Ref_Std = inputdata['TI_error_Ane2_Ref'].std()
    TI_diff_Ane2_Ref_Avg = inputdata['TI_diff_Ane2_Ref'].mean()
    TI_diff_Ane2_Ref_Std = inputdata['TI_diff_Ane2_Ref'].std()
    results['Ane2_Ref'] = [TI_error_Ane2_Ref_Avg,TI_error_Ane2_Ref_Std,TI_diff_Ane2_Ref_Avg,TI_diff_Ane2_Ref_Std]
    results.index = ['TI_error_mean', 'TI_error_std', 'TI_diff_mean', 'TI_diff_std']

    return results.T # T(ranspose) so that reporting looks good. 

def get_description_stats(inputdata):
    totalstats = get_stats_inBin(inputdata,1.75,20)
    belownominal = get_stats_inBin(inputdata,1.75,11.5)
    abovenominal = get_stats_inBin(inputdata,10,20)
    return totalstats, belownominal , abovenominal 

def get_representative_TI_15mps(inputdata):
    #this is the represetative TI, this is currently only done at a 1m/s bins not sure if this needs to be on .5m/s
    # TODO: find out if this needs to be on the 1m/s bin or the .5m/s bin 
    inputdata_TI15 = inputdata[inputdata['bins']==15]
    listofcols = ['RSD_TI','Ref_TI','Ane2_TI']
    if 'corrTI_RSD_WS' in inputdata.columns:
        listofcols.append('corrTI_RSD_WS')
    if 'corrWS_RSD_WS' in inputdata.columns:
        listofcols.append('corrWS_RSD_WS')
    results = inputdata_TI15[listofcols].describe()
    results.loc['Rep_TI',:]=results.loc['mean']+1.28*results.loc['std']
    results = results.loc[['mean','std','Rep_TI'],:].T
    return results

def write_resultstofile(df,ws, r_start,c_start):
    # write the regression results to file.
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, r_start):
        for c_idx, value in enumerate(row, c_start):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
def write_all_resultstofile(reg_results, TI_MBE_j_,TI_Diff_j_, rep_TI_results, TIbybin, count, total_stats, filename):
    wb = Workbook()
    ws = wb.active
    # regression results
    write_resultstofile(reg_results,ws,1,1)
    rownumber = 8
    # writing the total of all bins 
    totalcount = count[0].sum().sum()
    ws.cell(row=rownumber, column=1, value='Total Count')
    ws.cell(row=rownumber, column=2, value=totalcount)
    rownumber+=3

    for c in count:
        write_resultstofile(c,ws,rownumber,1)
        rownumber+=6

    for val in TI_MBE_j_:
        for val1 in val:
            try:
                write_resultstofile(val1,ws,rownumber,1)
                rownumber+=7
            except:
                print('Could not write a row in TI_MBE_j_, TI Corrected or WS corrected windspeeds are not provided so not writing them')
                
    for val in TI_Diff_j_:
            for val1 in val:
                try:
                    write_resultstofile(val1,ws,rownumber,1)
                    rownumber+=7
                except:
                    print('Could not write a row in TI_Diff_j_, TI Corrected or WS corrected windspeeds are not provided so not writing them')
    rownumber+=6
    for val in TIbybin:
        for i in val:
            try:
                write_resultstofile(i,ws,rownumber,1)
                rownumber+=6
            except:
                print('No data to write in one of the dataframes of TI by bin')
    rownumber+=6                        
    write_resultstofile(rep_TI_results, ws, rownumber,1)
    rownumber+=7
    headers_stats = dict(zip([0,1,2],['Total Bin Stats','Below nominal Stats','Above nominal Status']))
    for idx, val in enumerate(total_stats):
        ws.cell(row=rownumber, column=1, value=headers_stats[idx])
        rownumber+=1        
        write_resultstofile(val, ws, rownumber,1)        
        rownumber+=7
                
    wb.save(filename)

def get_inputfiles():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_filename", help="print this requires the input filename")
    parser.add_argument("config_file", help="this reuquires the excel configuration file")
    parser.add_argument("results_file", help="this reuquires the excel results file")
    args = parser.parse_args()
    print 'the input file is {}'.format(args.input_filename)
    print 'the input file is {}'.format(args.config_file)
    return args.input_filename, args.config_file, args.results_file
    
if __name__ == '__main__':
    input_filename, config_file, results_filename = get_inputfiles()
    inputdata = get_inputdata(input_filename, config_file)
    reg_results =get_ws_regression(inputdata)
    TI_MBE_j_,TI_Diff_j_ = get_TI_MBE_Diff_j(inputdata)
    rep_TI_results = get_representative_TI_15mps(inputdata)
    TIbybin = get_TI_bybin(inputdata)
    count = get_count_per_WSbin(inputdata,'RSD_WS')    
    total_stats = get_description_stats(inputdata)
    write_all_resultstofile(reg_results, TI_MBE_j_,TI_Diff_j_, rep_TI_results, TIbybin, count, total_stats, results_filename)
