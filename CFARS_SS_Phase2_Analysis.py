"""
This is the main script which is trying to clean and input data in the correct order into the internal input sheet. 
this will be used for the Phase 2 test summrization. 
Author : Nikhil Kondabala
Date: 11/16/2018 14:19
Updated for Phase 2 by Alexandra E. Arntsen and Barrett T. Goudeau from NRG Systems
Date: 8/01/2019 

Notes:

    -Ensure that data is pre-filtered for missing values before entering into dataset. Right now, data is filtered
    for WindCube missing value indicator (9999) so if data looks wonky, replace missing values with that. Note, if data
    is missing for any column of data all datapoints at that time will be removed from the analysis

    -Replaced Andrew's old ZX correction method with newer one, if you would like to use this one still renamed function
    as "perform_zx_correction_oldAB"

ToDo:
"""

import pandas as pd
import numpy as np
import statsmodels.formula.api as sm
import statsmodels.api as sm_api
import sys
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import argparse
import re
from sklearn import linear_model
from sklearn.metrics import mean_squared_error, r2_score
import warnings

def set_inputdataformat(config_file):
    # this function uses the input data function given in a configration file and converts into a dict 
    # of renamed cols which will be used to change the format from your internal format to a CFARS specific one 
    #  so that the rest of the script can work 
    df = pd.read_excel(config_file, usecols = [0,1]).dropna()

    df = df[((df['Header_YourData'] != 'RSD_model') &
             (df['Header_YourData'] != 'height_meters') &
             [re.search("correction",val) is None for val in df.Header_YourData])]

    # Run a quick check to make sure program exits gracefully if user makes a mistake in config
    #  (i.e, not having necessary variables, duplicate variables etc.)
    intColList = df.Header_YourData.tolist()
    cfarsColList = df.Header_CFARS_Python.tolist()

    if len(intColList) != len(set(intColList)):
        sys.exit('Looks like you have duplicate variables in the "Header_YourData" portion of the table, please correct and run again')

    if len(cfarsColList) != len(set(cfarsColList)):
        sys.exit('Looks like you have duplicate variables in the "Header_CFARS_Python" portion of the table, please correct and run again')

    # Run another quick check to ensure data fields that are necessary for analysis are entered. We MUST have reference
    requiredData = ['Ref_TI', 'Ref_WS', 'Ref_SD','Timestamp']
    if (set(requiredData).issubset(set(cfarsColList))) == False:
        missing = set(requiredData).difference(set(cfarsColList))
        sys.exit('You are missing the following variables in the Header_CFARS_Python that are necessary:\n' + str(missing)+
                 '\n Please fix and restart to run')
    # Check to see if we have an RSD to compare with 
    requiredData = ['RSD_TI','Ref_TI','RSD_WS','Ref_WS','Timestamp','RSD_SD','Ref_SD']
    if (set(requiredData).issubset(set(cfarsColList))) == False:
        missing = set(requiredData).difference(set(cfarsColList))
        print ('You are unable to apply all RSD correction methods, skipping RSD corrections due to missing:\n' + str(missing)+
                 '\n Please fix in order to run correction methods')
        requiredData = ['Ane2_TI', 'Ane2_WS', 'Ane2_SD']
        if (set(requiredData).issubset(set(cfarsColList))) == False:
            missing = set(requiredData).difference(set(cfarsColList))
            sys.exit('You are missing: ' + str(missing) + 'to compare to the reference instead of RSD.\n' +
                 '\n Please fix and restart to run')
    return dict(zip(intColList, cfarsColList))

def get_phaseii_metadata(config_file):
    df = pd.read_excel(config_file, usecols = [3,4]).dropna()
    try:
        model = df.Selection[df['Site Metadata'] == 'RSD Type:'].values[0]
    except:
        print('No Model Listed. Model coded as "unknown"')
        model = "unknown"

    try:
        height = df.Selection[df['Site Metadata'] == 'Comparison Height (m):'].values[0]
    except:
        print('No height listed. Height coded as "unknown"')
        height = "unknown"

    return model, height

def get_SiteMetadata(config_file):
    siteMetadata = pd.read_excel(config_file, usecols = [3,4,5],nrows=12)
    return(siteMetadata)

def get_FilteringMetadata(config_file):
    configMetadata = pd.read_excel(config_file, usecols = [3,4,5],nrows=8)
    return(configMetadata)

def check_for_corrections(config_file):
    apply_correction = True
    colLabels = pd.read_excel(config_file, usecols = [0,1])
    colLabels = list(colLabels.dropna()['Header_CFARS_Python'])
    rsd_cols = [s for s in colLabels if 'RSD' in s]
    requiredData = ['RSD_TI','RSD_WS','RSD_SD']
    if (set(requiredData).issubset(set(rsd_cols))) == False:
        apply_correction = False
    return(apply_correction)

def get_inputdata(filename, config_file):
    # This is the function to get the input data to analyze
    # the input data can be internal structure and does not need to be a set format. 
    #  the formatting information of the data is provided in the config file, the config file template is in the git hub repo

    if filename.split('.')[-1] == 'csv':
        inputdata = pd.read_csv(filename)  
    elif filename.split('.')[-1] == 'xlsx':
        inputdata = pd.read_excel(filename)
    else:
        print('Unkown input file type for the input data , please consider changing it to csv')
        sys.exit()

    try:
        rename_cols  = set_inputdataformat(config_file)
    except Exception as e:
        print('There is an error in the configuration file')
        sys.exit()

    inputdata = inputdata.rename(index=str,columns=rename_cols)
    inputdata = inputdata.dropna()
    # this needs to be thought through, in essence we are applying a wholesale drop for non exisiting data
    inputdata['bins']=inputdata['Ref_WS'].round(0) # this acts as bin because the bin defination is between the two half integer values

    bins_p5_interval= pd.interval_range(start=.25,end=20,freq=.5, closed='left')# this is creating a interval range of .5 starting at .25
    out =pd.cut(x= inputdata['Ref_WS'], bins=bins_p5_interval)
    inputdata['bins_p5']=out.apply(lambda x: x.mid) # the middle of the interval is used as a catagorical label 
    inputdata = inputdata[inputdata['Ref_TI']!=0] # we can only analyze where the ref_TI is not 0

    #filter data, look for missing values *** Need to add more to this list as we see fit ***
    inputdata = inputdata.drop('Timestamp',1).replace(9999, np.NaN)
    inputdata = inputdata.dropna()

    return inputdata

#TODO: export the representative TI plot
# def plot_rep_TI(representative_TI, projectname):
#     # repcols = [u'RSD_TI_rep',u'Ref_TI_rep', u'Ane2_TI_rep']
#     # plt = representative_TI[repcols].plot(title='Rep TI plot, TI_mean_bin+TI_sd_bin*1.28', ylable='Rep TI')
#     # plt.figure.savefig('./{}_rep_TIplot.png'.format(projectname))

def get_regression_sm(x,y):
    # get the linear least squares fit, this uses statsmodel which needs to be installed outside of Anaconda
    x = sm_api.add_constant(x)
    model = sm_api.OLS(y,x,missing='drop').fit()
    result = model.params
    result['Rsquared'] = model.rsquared
    result['WSdiff'] = abs((x.iloc[:,1]-y).mean())
    return [result[1], result[0], result[2], result[3]]

def get_regression(x,y):
    # get the linear least squaes fit, this uses the sklearn model which is in Anaconda already.  
    lm = linear_model.LinearRegression()
    lm.fit(x.to_frame(),y.to_frame())
    result = [lm.coef_[0][0],lm.intercept_[0]]
    result.append(lm.score(x.to_frame(),y.to_frame()))
    result.append(abs((x-y).mean()))
    return result

def get_modelRegression(inputdata, column1, column2):
    '''
    :param inputdata: input data (dataframe)
    :param column1: string, column name for x-variable
    :param column2: string, column name for y-variable
    :param columnNameOut: string, column name for predicted value
    :return: dict with out of regression
    '''
    x = inputdata[column1].values.reshape(len(inputdata[column1]),1)
    y = inputdata[column2].values.reshape(len(inputdata[column2]),1)

    regr = linear_model.LinearRegression()
    regr.fit(x, y)
    slope = regr.coef_[0]
    intercept = regr.intercept_
    predict = regr.predict(x)
    r = np.corrcoef(x, y)[0, 1]
    r2 = r2_score(x, predict)  # coefficient of determination, explained variance
    mse = mean_squared_error(x, predict, multioutput='raw_values')

    rmse = np.sqrt(mse)
    results = {'c':intercept,'m':slope,'r':r,'r2': r2, 'mse': mse, 'rmse': rmse, 'predicted':predict}
    return results

def get_ws_regression(inputdata):
    # get the ws regression results for all the col required pairs. 
    results = pd.DataFrame(columns=['m','c','rsquared','ws_diff'])

    if 'RSD_WS' in inputdata.columns:
        results_rsd_ref = get_regression(inputdata['RSD_WS'],inputdata['Ref_WS']) # this is to get the rsd and ref anemometer results. 
        results.loc['WS_regression_RSD_Ref'] = results_rsd_ref
    else:
        results.loc['WS_regression_RSD_Ref'] = [None, None, None, None]
    if 'corrTI_RSD_WS' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results. 
        results_corrTI_rsd_ref = get_regression(inputdata['corrTI_RSD_WS'],inputdata['Ref_WS'])
        results.loc['WS_regression_corrTI_RSD_Ref']=results_corrTI_rsd_ref
    else:
        results.loc['WS_regression_corrTI_RSD_Ref']=[None, None, None, None]
    if 'corrWS_RSD_WS' in inputdata.columns:
        results_corrWS_rsd_ref = get_regression(inputdata['corrTI_RSD_WS'],inputdata['Ref_WS'])
        results.loc['WS_regression_corrWS_RSD_Ref']=results_corrWS_rsd_ref
    else:
        results.loc['WS_regression_corrWS_RSD_Ref']=[None, None, None, None]
    
    if 'Ane2_WS' in inputdata.columns:
        results_Ane2_ref = get_regression(inputdata['Ane2_WS'],inputdata['Ref_WS'])
        results.loc['WS_regression_Ane2_Ref']=results_Ane2_ref
    else:
        results.loc['WS_regression_Ane2_Ref']=[None, None, None, None]
        
    return results

def perform_ge_correction(inputdata):
    import matplotlib.pyplot as plt

    # filter out RSD_TI > 0.3
    # model = get_regression(filtered RSD_TI, Ref_TI)
    # apply model to RSD_TI
    # check if corrTI_RSD_TI is present. If so, throw an error
    # inputdata['corrTI_RSD_TI'] = model output

    filtered_Ref_TI = inputdata['Ref_TI'][inputdata['RSD_TI']<0.3]
    filtered_RSD_TI = inputdata['RSD_TI'][inputdata['RSD_TI']<0.3]


    model = get_regression(filtered_RSD_TI, filtered_Ref_TI)
    RSD_TI = inputdata['RSD_TI'].copy()
    RSD_TI[RSD_TI>=0.3] = np.nan
    RSD_TI[RSD_TI<0.3] = model[0]*RSD_TI[RSD_TI<0.3]+model[1]
    inputdata['corrTI_RSD_TI'] = RSD_TI


    results = pd.DataFrame(columns=['sensor','height','correction','m',\
                                    'c','rsquared','ws_diff'])

    results.loc['TI_regression_corrTI_RSD_Ref',3:7] = model
    return inputdata, results


def perform_eon_correction(inputdata):
    # determine sigma, mean WS from RSD Timeseries (10 minutes)?
    # determine which wind speed bin data from 10 min avg falls into, apply correction using slope/offset
    # subtract bias from the now corrected RSD for each speed range
    # Calculate TI using corrected sigmaRSD

    inputdata['CorrectedSD_Eon'] = inputdata['RSD_SD']

    for i in inputdata.index:
        if inputdata.at[i,'RSD_WS'] <4:
            inputdata.at[i,'CorrectedSD_Eon'] = inputdata.at[i,'CorrectedSD_Eon']
        elif inputdata.at[i,'RSD_WS'] >= 4 and inputdata.at[i,'RSD_WS'] <8:
            sigW = inputdata.at[i, 'CorrectedSD_Eon']
            inputdata.at[i, 'CorrectedSD_Eon'] = (sigW * 1.116763 + .024685) - (.029 *(sigW * 1.116763 + .024685))
        elif inputdata.at[i,'RSD_WS'] >= 8 and inputdata.at[i,'RSD_WS'] <12:
            sigW = inputdata.at[i, 'CorrectedSD_Eon']
            inputdata.at[i, 'CorrectedSD_Eon'] = (sigW * 1.064564 + .040596) - (-.161 *(sigW * 1.064564 + .040596))
        elif inputdata.at[i, 'RSD_WS'] >= 12 and inputdata.at[i, 'RSD_WS'] < 16:
            sigW = inputdata.at[i, 'CorrectedSD_Eon']
            inputdata.at[i, 'CorrectedSD_Eon'] = (sigW * .97865 + .124371) - (-.093 * (sigW * .97865 + .124371))
        else:
            inputdata.at[i,'CorrectedSD_Eon'] = inputdata.at[i,'CorrectedSD_Eon']
    inputdata['corrTI_RSD_TI'] = inputdata['CorrectedSD_Eon']/inputdata['RSD_WS']

    return inputdata

def zx_correct(x):
    return(1./np.interp(x,[20,90],[1.037,0.918],left=1.037,right=0.918))

def perform_zx_correction_oldAB(inputdata, sta):
    if 'corrTI_RSD_TI' in inputdata.columns:
        sys.exit("Corrected TI already present in data. \
                 \nNo correction performed. Terminating process...")
    else:
        RSD_TI = inputdata['RSD_TI'].copy()
        RSD_TI = RSD_TI * zx_correct(sta)
        inputdata['corrTI_RSD_TI'] = RSD_TI

    return inputdata

def perform_zx_correction(inputdata, height):

    RSD_TI = inputdata['RSD_TI'].copy()

    # Derived from ZX_Lidars_TI_Correction_Method_summary.pdf, pg7
    print (height)
    slope = .1518 * height +21.311
    intercept = .3831 * height +37.229

    ti = RSD_TI * (1-((slope * np.log10(RSD_TI))/(100)) - (intercept/100))

    inputdata['corrTI_RSD_TI'] = ti
    return inputdata


def get_representative_TI(inputdata):
    # get the representaive TI
    representative_TI = inputdata[['RSD_TI','Ref_TI','Ane2_TI','bins']].groupby(by=['bins']).agg(['mean','std',lambda x: x.mean()+1.28*x.std()])
    representative_TI.columns = ['RSD_TI_mean','RSD_TI_std','RSD_TI_rep','Ref_TI_mean','Ref_TI_std','Ref_TI_rep','Ane2_TI_mean','Ane2_TI_std','Ane2_TI_rep']
    return representative_TI

def get_count_per_WSbin(inputdata, column):
    # Count per wind speed bin
    inputdata = inputdata[(inputdata['bins_p5'].astype(float)>1.5) & (inputdata['bins_p5'].astype(float)<21)]  
    resultsstats_bin = inputdata[[column,'bins']].groupby(by = 'bins').agg(['count'])
    resultsstats_bin_p5 = inputdata[[column,'bins_p5']].groupby(by = 'bins_p5').agg(['count'])
    resultsstats_bin = pd.DataFrame(resultsstats_bin.unstack()).T
    resultsstats_bin.index = [column]
    resultsstats_bin_p5 = pd.DataFrame(resultsstats_bin_p5.unstack()).T
    resultsstats_bin_p5.index = [column]   
    return resultsstats_bin, resultsstats_bin_p5

def get_stats_per_WSbin(inputdata, column):
    # this will be used as a base function for all frequency agg caliculaitons for each bin to get the stats per wind speed bins
    inputdata = inputdata[(inputdata['bins_p5'].astype(float)>1.5) & (inputdata['bins_p5'].astype(float)<21)]  
    resultsstats_bin = inputdata[[column,'bins']].groupby(by = 'bins').agg(['mean','std'])
    resultsstats_bin_p5 = inputdata[[column,'bins_p5']].groupby(by = 'bins_p5').agg(['mean','std'])
    resultsstats_bin = pd.DataFrame(resultsstats_bin.unstack()).T
    resultsstats_bin.index = [column]
    resultsstats_bin_p5 = pd.DataFrame(resultsstats_bin_p5.unstack()).T
    resultsstats_bin_p5.index = [column]
    return resultsstats_bin, resultsstats_bin_p5

def get_RMSE_per_WSbin(inputdata,column):
    """
    get RMSE with no fit model, just based on residual being the reference
    """
    squared_TI_Diff_j_RSD_Ref, squared_TI_Diff_jp5_RSD_Ref = get_stats_per_WSbin(inputdata,column)
    TI_RMSE_j = squared_TI_Diff_j_RSD_Ref**(.5)
    TI_RMSE_jp5 = squared_TI_Diff_jp5_RSD_Ref**(.5)

    return TI_RMSE_j, TI_RMSE_jp5

def get_TI_MBE_Diff_j(inputdata):
    
    TI_MBE_j_ = []
    TI_Diff_j_ = []
    TI_RMSE_j_ = []
    
    # get the bin wise stats for both diff and error between RSD and Ref
    if 'RSD_TI' in inputdata.columns:
        inputdata['TI_diff_RSD_Ref'] = inputdata['RSD_TI'] - inputdata[
            'Ref_TI']  # caliculating the diff in ti for each timestamp
        inputdata['TI_error_RSD_Ref'] = inputdata['TI_diff_RSD_Ref'] / inputdata[
            'Ref_TI']  # calculating the error for each timestamp
        inputdata['TI_RMSE_RSD_Ref'] = inputdata['TI_diff_RSD_Ref'] * inputdata[
            'TI_diff_RSD_Ref']  # calculating squared diff each Time
        TI_MBE_j_RSD_Ref, TI_MBE_jp5_RSD_Ref = get_stats_per_WSbin(inputdata, 'TI_error_RSD_Ref')
        TI_Diff_j_RSD_Ref, TI_Diff_jp5_RSD_Ref = get_stats_per_WSbin(inputdata, 'TI_diff_RSD_Ref')
        TI_RMSE_j_RSD_Ref, TI_RMSE_jp5_RSD_Ref = get_RMSE_per_WSbin(inputdata, 'TI_RMSE_RSD_Ref')

        TI_MBE_j_.append([TI_MBE_j_RSD_Ref, TI_MBE_jp5_RSD_Ref])
        TI_Diff_j_.append([TI_Diff_j_RSD_Ref, TI_Diff_jp5_RSD_Ref])
        TI_RMSE_j_.append([TI_RMSE_j_RSD_Ref, TI_RMSE_jp5_RSD_Ref])
    else:
        TI_MBE_j_.append(pd.DataFrame([None, None]))
        TI_Diff_j_.append(pd.DataFrame([None, None]))
        TI_RMSE_j_.append(pd.DataFrame([None, None]))

    # get the bin wise stats for both diff and error and RMSE between RSD corrected for TI and Ref

    if 'corrTI_RSD_TI' in inputdata.columns:  # this is checking if corrected TI windspeed is present in the input data and using that for getting the results.
        inputdata['TI_diff_corrTI_RSD_Ref'] = inputdata['corrTI_RSD_TI'] - inputdata[
            'Ref_TI']  # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrTI_RSD_Ref'] = inputdata['TI_diff_corrTI_RSD_Ref'] / inputdata[
            'Ref_TI']  # calculating the error for each timestamp
        inputdata['TI_RMSE_corrTI_RSD_Ref'] = inputdata['TI_diff_corrTI_RSD_Ref'] * inputdata[
            'TI_diff_corrTI_RSD_Ref']
        TI_MBE_j_corrTI_rsd_ref, TI_MBE_jp5_corrTI_rsd_ref = get_stats_per_WSbin(inputdata, 'TI_error_corrTI_RSD_Ref')
        TI_Diff_j_corrTI_rsd_ref, TI_Diff_jp5_corrTI_rsd_ref = get_stats_per_WSbin(inputdata, 'TI_diff_corrTI_RSD_Ref')
        TI_RMSE_j_corrTI_rsd_ref, TI_RMSE_jp5_corrTI_rsd_ref = get_RMSE_per_WSbin(inputdata,'TI_RMSE_corrTI_RSD_Ref')
        
        TI_MBE_j_.append([TI_MBE_j_corrTI_rsd_ref, TI_MBE_jp5_corrTI_rsd_ref])
        TI_Diff_j_.append([TI_Diff_j_corrTI_rsd_ref, TI_Diff_jp5_corrTI_rsd_ref])
        TI_RMSE_j_.append([TI_RMSE_j_corrTI_rsd_ref, TI_RMSE_jp5_corrTI_rsd_ref])
    else:
        TI_MBE_j_.append(pd.DataFrame([None, None]))
        TI_Diff_j_.append(pd.DataFrame([None, None]))
        TI_RMSE_j_.append(pd.DataFrame([None, None]))

    # get the bin wise stats for both diff and error between RSD corrected for ws and Ref

    if 'corrWS_RSD_TI' in inputdata.columns:  # this is checking if the corrected WS method was used for the wind speed and TI
        inputdata['TI_diff_corrWS_RSD_Ref'] = inputdata['corrWS_RSD_TI'] - inputdata[
            'Ref_TI']  # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrWS_RSD_Ref'] = inputdata['TI_diff_corrWS_RSD_Ref'] / inputdata[
            'Ref_TI']  # calculating the error for each timestamp
        inputdata['TI_RMSE_corrWS_RSD_Ref'] = inputdata['TI_diff_corrWS_RSD_Ref'] * inputdata[
            'TI_diff_corrWS_RSD_Ref']
        TI_MBE_j_corrWS_rsd_ref, TI_MBE_jp5_corrWS_rsd_ref = get_stats_per_WSbin(inputdata, 'TI_error_corrWS_RSD_Ref')
        TI_Diff_j_corrWS_rsd_ref, TI_Diff_jp5_corrWS_rsd_ref = get_stats_per_WSbin(inputdata, 'TI_diff_corrWS_RSD_Ref')
        TI_RMSE_j_corrWS_rsd_ref, TI_RMSE_jp5_corrWS_rsd_ref = get_RMSE_per_WSbin(inputdata,'TI_RMSE_corrWS_RSD_Ref')
        TI_MBE_j_.append([TI_MBE_j_corrWS_rsd_ref, TI_MBE_jp5_corrWS_rsd_ref])
        TI_Diff_j_.append([TI_Diff_j_corrWS_rsd_ref, TI_Diff_jp5_corrWS_rsd_ref])
        TI_RMSE_j_.append([TI_RMSE_j_corrWS_rsd_ref, TI_RMSE_jp5_corrWS_rsd_ref])
    else:
        TI_MBE_j_.append(pd.DataFrame([None, None]))
        TI_Diff_j_.append(pd.DataFrame([None, None]))
        TI_RMSE_j_.append(pd.DataFrame([None, None]))

    # get bin wise stats for the ane2 and ref anemometer
    if 'Ane2_TI' in inputdata.columns:
        inputdata['TI_diff_Ane2_Ref'] = inputdata['Ane2_TI'] - inputdata[
            'Ref_TI']  # caliculating the diff in ti for each timestamp
        inputdata['TI_error_Ane2_Ref'] = inputdata['TI_diff_Ane2_Ref'] / inputdata[
            'Ref_TI']  # calculating the error for each timestamp
        inputdata['TI_RMSE_Ane2_Ref'] = inputdata['TI_diff_Ane2_Ref'] * inputdata['TI_diff_Ane2_Ref']
        TI_MBE_j_Ane2_Ref, TI_MBE_jp5_Ane2_Ref = get_stats_per_WSbin(inputdata, 'TI_error_Ane2_Ref')
        TI_Diff_j_Ane2_Ref, TI_Diff_jp5_Ane2_Ref = get_stats_per_WSbin(inputdata, 'TI_diff_Ane2_Ref')
        TI_RMSE_j_Ane2_ref, TI_RMSE_jp5_Ane2_ref = get_RMSE_per_WSbin(inputdata, 'TI_RMSE_Ane2_Ref')
        TI_MBE_j_.append([TI_MBE_j_Ane2_Ref, TI_MBE_jp5_Ane2_Ref])
        TI_Diff_j_.append([TI_Diff_j_Ane2_Ref, TI_Diff_jp5_Ane2_Ref])
        TI_RMSE_j_.append([TI_RMSE_j_Ane2_ref, TI_RMSE_jp5_Ane2_ref])
    else:
        TI_MBE_j_.append([None, None])
        TI_Diff_j_.append([None, None])
        TI_RMSE_j_.append([None, None])

    return TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_


def get_TI_bybin(inputdata):
    results = []

    if 'RSD_TI' in inputdata.columns:
        RSD_TI_j, RSD_TI_jp5 = get_stats_per_WSbin(inputdata,'RSD_TI')
        results.append([RSD_TI_j, RSD_TI_jp5])
    else:
        results.append([None, None])
        
    Ref_TI_j, Ref_TI_jp5 = get_stats_per_WSbin(inputdata,'Ref_TI')
    results.append([Ref_TI_j, Ref_TI_jp5])

    if 'corrTI_RSD_TI' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results. 
        corrTI_RSD_TI_j, corrTI_RSD_TI_jp5 = get_stats_per_WSbin(inputdata,'corrTI_RSD_TI')
        results.append([corrTI_RSD_TI_j,corrTI_RSD_TI_jp5])
    else:
        results.append(pd.DataFrame([None, None]))

    # get the bin wise stats for both diff and error between RSD corrected for ws and Ref
        
    if 'corrWS_RSD_TI' in inputdata.columns: #this is checking if the corrected WS method was used for the wind speed and TI
        corrWS_RSD_TI_j, corrWS_RSD_TI_jp5 = get_stats_per_WSbin(inputdata,'corrWS_RSD_TI')
        results.append([corrWS_RSD_TI_j, corrWS_RSD_TI_jp5])
    else:
        results.append(pd.DataFrame([None, None]))

    if 'Ane2_TI' in inputdata.columns:
        Ane2_TI_j, Ane2_TI_jp5 = get_stats_per_WSbin(inputdata,'Ane2_TI')
        results.append([Ane2_TI_j, Ane2_TI_jp5])
    else:
        results.append(pd.DataFrame([None, None]))
        
    return results

def get_stats_inBin(inputdata_m,start,end):    
    # this was discussed in the meeting , but the results template didn't ask for this. 
    inputdata = inputdata_m.loc[(inputdata_m['Ref_WS']>start) & (inputdata_m['Ref_WS']<=end)].copy()

    if 'RSD_TI' in inputdata.columns:
        inputdata['TI_diff_RSD_Ref'] = inputdata['RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_RSD_Ref'] = inputdata['TI_diff_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp

    # Make sure nans are dropped
    inputdata= inputdata.dropna()

    if 'RSD_TI' in inputdata.columns:
        TI_error_RSD_Ref_Avg = inputdata['TI_error_RSD_Ref'].mean()
        TI_error_RSD_Ref_Std = inputdata['TI_error_RSD_Ref'].std()
        TI_diff_RSD_Ref_Avg = inputdata['TI_diff_RSD_Ref'].mean()
        TI_diff_RSD_Ref_Std = inputdata['TI_diff_RSD_Ref'].std()
    else:
        TI_error_RSD_Ref_Avg = None
        TI_error_RSD_Ref_Std = None
        TI_diff_RSD_Ref_Avg = None
        TI_diff_RSD_Ref_Std = None

    #RSD V Reference
    if 'RSD_TI' in inputdata.columns:
        modelResults = get_modelRegression(inputdata, 'RSD_TI', 'Ref_TI')
        rmse = modelResults['rmse'][0]
        slope = modelResults['m'][0]
        offset = modelResults['c'][0]
        # print('rmse: ', rmse)
        # print('slope: ', slope)
        # print('offset: ',offset )
    else:
        rmse = None
        slope = None
        offset = None

    results = pd.DataFrame([TI_error_RSD_Ref_Avg,TI_error_RSD_Ref_Std,TI_diff_RSD_Ref_Avg,TI_diff_RSD_Ref_Std,slope, offset,rmse],columns=['RSD_Ref'])
    
    if 'corrTI_RSD_TI' in inputdata.columns: # this is checking if corrected TI windspeed is present in the input data and using that for getting the results.
        #Cor RSD vs Reg RSD
        inputdata['TI_diff_corrTI_RSD_Ref']= inputdata['corrTI_RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrTI_RSD_Ref']= inputdata['TI_diff_corrTI_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_error_corrTI_RSD_Ref_Avg = inputdata['TI_error_corrTI_RSD_Ref'].mean()
        TI_error_corrTI_RSD_Ref_Std = inputdata['TI_error_corrTI_RSD_Ref'].std()
        TI_diff_corrTI_RSD_Ref_Avg = inputdata['TI_diff_corrTI_RSD_Ref'].mean()
        TI_diff_corrTI_RSD_Ref_Std = inputdata['TI_diff_corrTI_RSD_Ref'].std()


        modelResults = get_modelRegression(inputdata, 'corrTI_RSD_TI', 'Ref_TI')

        rmse = modelResults['rmse'][0]
        slope = modelResults['m'][0]
        offset = modelResults['c'][0]

        results['CorrTI_RSD_Ref'] = [TI_error_corrTI_RSD_Ref_Avg,TI_error_corrTI_RSD_Ref_Std, TI_diff_corrTI_RSD_Ref_Avg, TI_diff_corrTI_RSD_Ref_Std,slope, offset,rmse]
    else:
        results['CorrTI_RSD_Ref'] = [None, None, None, None, None, None, None]
        
    # get the bin wise stats for both diff and error between RSD corrected for ws and Ref
        
    if 'corrWS_RSD_TI' in inputdata.columns: #this is checking if the corrected WS method was used for the wind speed and TI
        #corr ws rsd vs ws ref
        inputdata['TI_diff_corrWS_RSD_Ref']= inputdata['corrWS_RSD_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_corrWS_RSD_Ref']= inputdata['TI_diff_corrWS_RSD_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_error_corrWS_RSD_Ref_Avg = inputdata['TI_error_corrWS_RSD_Ref'].mean()
        TI_error_corrWS_RSD_Ref_Std = inputdata['TI_error_corrWS_RSD_Ref'].std()
        TI_diff_corrWS_RSD_Ref_Avg = inputdata['TI_diff_corrWS_RSD_Ref'].mean()
        TI_diff_corrWS_RSD_Ref_Std = inputdata['TI_diff_corrWS_RSD_Ref'].std()
        rmse = np.nan
        slope = np.nan
        offset = np.nan
        results['corrWS_RSD_Ref'] = [TI_error_corrWS_RSD_Ref_Avg,TI_error_corrWS_RSD_Ref_Std,TI_diff_corrWS_RSD_Ref_Avg,TI_diff_corrWS_RSD_Ref_Std,slope, offset,rmse]
    else:
        results['corrWS_RSD_Ref'] =[None, None, None, None, None, None, None]

    #anem 2 vs ref

    if 'Ane2_TI' in inputdata.columns:
        inputdata['TI_diff_Ane2_Ref']= inputdata['Ane2_TI']-inputdata['Ref_TI'] # caliculating the diff in ti for each timestamp
        inputdata['TI_error_Ane2_Ref']= inputdata['TI_diff_Ane2_Ref']/inputdata['Ref_TI'] #calculating the error for each timestamp
        TI_error_Ane2_Ref_Avg = inputdata['TI_error_Ane2_Ref'].mean()
        TI_error_Ane2_Ref_Std = inputdata['TI_error_Ane2_Ref'].std()
        TI_diff_Ane2_Ref_Avg = inputdata['TI_diff_Ane2_Ref'].mean()
        TI_diff_Ane2_Ref_Std = inputdata['TI_diff_Ane2_Ref'].std()
        results['Ane2_Ref'] = [TI_error_Ane2_Ref_Avg,TI_error_Ane2_Ref_Std,TI_diff_Ane2_Ref_Avg,TI_diff_Ane2_Ref_Std,slope, offset,rmse]
    else:
        results['Ane2_Ref'] = [None,None,None,None,None, None, None]

    
    results.index = ['TI_error_mean', 'TI_error_std', 'TI_diff_mean', 'TI_diff_std', 'Slope','Offset','RMSE']

    return results.T # T(ranspose) so that reporting looks good. 

def get_description_stats(inputdata):
    totalstats = get_stats_inBin(inputdata,1.75,20)
    belownominal = get_stats_inBin(inputdata,1.75,11.5)
    abovenominal = get_stats_inBin(inputdata,10,20)
    return totalstats, belownominal , abovenominal 

def get_representative_TI_15mps(inputdata):
    #this is the represetative TI, this is currently only done at a 1m/s bins not sure if this needs to be on .5m/s
    # TODO: find out if this needs to be on the 1m/s bin or the .5m/s bin 
    inputdata_TI15 = inputdata[inputdata['bins']==15]
    listofcols = ['Ref_TI']
    if 'RSD_TI' in inputdata.columns:
        listofcols.append('RSD_TI')
    if 'Ane2_TI' in inputdata.columns:
        listofcols.append('Ane2_TI')
    if 'corrTI_RSD_WS' in inputdata.columns:
        listofcols.append('corrTI_RSD_WS')
    if 'corrWS_RSD_WS' in inputdata.columns:
        listofcols.append('corrWS_RSD_WS')
    results = inputdata_TI15[listofcols].describe()
    results.loc['Rep_TI',:]=results.loc['mean']+1.28*results.loc['std']
    results = results.loc[['mean','std','Rep_TI'],:].T
    results.columns = ['mean_15mps','std_15mps', 'Rep_TI']
    return results

def write_resultstofile(df,ws,r_start,c_start):
    # write the regression results to file.
    rows = dataframe_to_rows(df)
    for r_idx, row in enumerate(rows, r_start):
        for c_idx, value in enumerate(row, c_start):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
def write_all_resultstofile(reg_results, allTIRegressResults, TI_MBE_jList,TI_Diff_jList, TI_RMSE_jList, rep_TI_resultsList,\
                            TIbybinList, countList, total_statsList, filename, \
                            corrTI_resultsList, correctionTag, siteMetadata, filterMetadata):

    wb = Workbook()
    ws = wb.active
    ws.title = correctionTag[0] + ' Corrections'
    
    for i in range(len(correctionTag)):
        sheetName = correctionTag[i] + ' Corrections'
        TI_MBE_j_ = TI_MBE_jList[i]
        TI_Diff_j_ = TI_Diff_jList[i]
        TI_RMSE_j_ = TI_RMSE_jList[i]
        rep_TI_results = rep_TI_resultsList[i]
        TIbybin = TIbybinList[i]
        count = countList[i]
        total_stats = total_statsList[i]
        corrTI_results = corrTI_resultsList[i]

        if i >0:
            ws = wb.create_sheet(title=sheetName)

        # regression results
        write_resultstofile(reg_results,ws,1,1)

        m = allTIRegressResults['m'][0]
        c = allTIRegressResults['c'][0]
        r2 = allTIRegressResults['r2']
        rmse = allTIRegressResults['rmse'][0]
        ws.cell(row = 8, column = 1, value = 'TI_RSDvsREF_NoCorrection')
        ws.cell(row = 8, column = 2, value = 'm')
        ws.cell(row = 9, column = 2, value = m)
        ws.cell(row = 8, column = 3, value = 'c')
        ws.cell(row = 9, column = 3, value = c)
        ws.cell(row = 8, column = 4, value = 'rsquared')
        ws.cell(row = 9, column = 4, value = r2)
        ws.cell(row = 8, column = 5, value = 'rmse')
        ws.cell(row = 9, column = 5, value = rmse)


        rownumber = 12
        # writing the total of all bins
        totalcount = count[0].sum().sum()
        ws.cell(row=rownumber, column=1, value='Total Count')
        ws.cell(row=rownumber, column=2, value=totalcount)
        rownumber+=2

        write_resultstofile(corrTI_results,ws,rownumber,1)
        rownumber+=5

        for c in count:
            write_resultstofile(c,ws,rownumber,1)
            rownumber+=6

        for val in TI_MBE_j_:
            for val1 in val:
                try:
                    write_resultstofile(val1,ws,rownumber,1)
                    rownumber+=7
                except:
                    pass
                    #print('Could not write a row in TI_MBE_j_, TI Corrected or WS corrected windspeeds are not provided so not writing them')
                    
        for val in TI_RMSE_j_:
            for val1 in val:
                try:
                    write_resultstofile(val1,ws,rownumber,1)
                    rownumber+=7
                except:
                    pass
                    #print('Could not write a row in TI_RMSE_j_, TI Corrected or WS corrected windspeeds are not provided so not writing them')

        for val in TI_Diff_j_:
                for val1 in val:
                    try:
                        write_resultstofile(val1,ws,rownumber,1)
                        rownumber+=7
                    except:
                        pass
                        #print('Could not write a row in TI_Diff_j_, TI Corrected or WS corrected windspeeds are not provided so not writing them')
        rownumber+=6
        for val in TIbybin:
            for i in val:
                try:
                    write_resultstofile(i,ws,rownumber,1)
                    rownumber+=6
                except:
                    print('No data to write in one of the dataframes of TI by bin')
        rownumber+=6
        write_resultstofile(rep_TI_results, ws, rownumber,1)
        rownumber+=8
        headers_stats = dict(zip([0,1,2],['Total Bin Stats','Below nominal Stats','Above nominal Status']))
        for idx, val in enumerate(total_stats):
            ws.cell(row=rownumber, column=1, value=headers_stats[idx])
            rownumber+=1
            write_resultstofile(val, ws, rownumber,1)
            rownumber+=7
    a = wb.create_sheet(title = 'Site Metadata')
    for r in dataframe_to_rows(siteMetadata, index = False):
        a.append(r)
    b = wb.create_sheet(title ='Filter Metadata')
    for r in dataframe_to_rows(filterMetadata, index = False):
        b.append(r)
    wb.save(filename)

def get_inputfiles():
    parser = argparse.ArgumentParser()
    parser.add_argument("input_filename", help="print this requires the input filename")
    parser.add_argument("config_file", help="this requires the excel configuration file")
    parser.add_argument("results_file", help="this requires the excel results file")
    args = parser.parse_args()
    print('the input file is {}'.format(args.input_filename))
    print('the input file is {}'.format(args.config_file))
    return args.input_filename, args.config_file, args.results_file
    
if __name__ == '__main__':
    input_filename, config_file, results_filename = get_inputfiles()
    apply_correction = check_for_corrections(config_file)
    siteMetadata = get_SiteMetadata(config_file)
    filterMetadata = get_FilteringMetadata(config_file)
    inputdata = get_inputdata(input_filename, config_file)
    reg_results = get_ws_regression(inputdata)

    #This line calculates TI regression for all data, without it code will  break when writing...
    if 'RSD_TI' in inputdata.columns:
        allTIRegressResults = get_modelRegression(inputdata, 'RSD_TI','Ref_TI')
    else:
        allTIRegressResults = {'c':[None],'m':[None],'r':[None],'r2': None, 'mse': [None], 'rmse': [None], 'predicted':[None]}
    sensor, height = get_phaseii_metadata(config_file)
    print(str(sensor)+", z = "+str(height))


    # Corrections
    if apply_correction == True:
        if siteMetadata.iloc[6]['Selection'] == 'Triton':
            correctionsList = ['VAISALA','GE', 'ZX','ZXOld','L-TERRA']
        else:
            correctionsList = ['GE', 'ZX','ZXOld','EON', 'L-TERRA']   
    else:
        correctionsList = ['']
        
    TI_MBEList = []
    TI_DiffList = []
    TI_RMSEList = []
    rep_TI_resultsList =[]
    TIBinList = []
    countList = []
    total_StatsList = []
    lm_CorrList = []
    correctionTag = []

    for i in correctionsList:
        if i == '':
            print ('No RSD correction applied')
            correctionTag.append('NoCorrection')

            lm_corr = pd.DataFrame(columns=['sensor', 'height', 'correction', 'm', \
                                            'c', 'rsquared'])
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'sensor'] = None
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'height'] = None
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'correction'] = 'No correction'
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'm'] = None
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'c'] = None
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'rsquared'] = None
            lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'rmse'] = None

            #Compute model regression results
            TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdata)
            rep_TI_results = get_representative_TI_15mps(inputdata)
            TIbybin = get_TI_bybin(inputdata)
            count = get_count_per_WSbin(inputdata, 'Ane2_WS')
            total_stats = get_description_stats(inputdata)

            TI_MBEList.append(TI_MBE_j_)
            TI_DiffList.append(TI_Diff_j_)
            TI_RMSEList.append(TI_RMSE_j_)
            rep_TI_resultsList.append(rep_TI_results)
            TIBinList.append(TIbybin)
            countList.append(count)
            total_StatsList.append(total_stats)
            lm_CorrList.append(lm_corr)
            
        if i == 'GE':
            if 'corrTI_RSD_TI' in inputdata.columns:
                print("Corrected TI already present in data. GE correction not performed.")
            else:
                inputdataGE, lm_corrGE = perform_ge_correction(inputdata.copy())
                print("GE: y = "+str(round(lm_corrGE.iloc[0,3],2))+"*x + "+\
                      str(round(lm_corrGE.iloc[0,4],2)))
                lm_corrGE['sensor'] = sensor
                lm_corrGE['height'] = height
                lm_corrGE['correction'] = 'GE'

                TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdataGE)
                rep_TI_results = get_representative_TI_15mps(inputdataGE)
                TIbybin = get_TI_bybin(inputdataGE)
                count = get_count_per_WSbin(inputdataGE, 'RSD_WS')
                total_stats = get_description_stats(inputdataGE)

                TI_MBEList.append(TI_MBE_j_)
                TI_DiffList.append(TI_Diff_j_)
                TI_RMSEList.append(TI_RMSE_j_)
                rep_TI_resultsList.append(rep_TI_results)
                TIBinList.append(TIbybin)
                countList.append(count)
                total_StatsList.append(total_stats)
                lm_CorrList.append(lm_corrGE)
                correctionTag.append(i)

        if i == 'EON':
            print('EON Correction Applied to Data')
            inputdataEON = perform_eon_correction(inputdata.copy())
            results = get_modelRegression(inputdataEON, 'corrTI_RSD_TI','Ref_TI')


            lm_corrEON = pd.DataFrame(columns=['sensor', 'height', 'correction', 'm', \
                                            'c', 'rsquared'])
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'sensor'] = sensor
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'height'] = height
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'correction'] = 'EON'
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'm'] = results['m'][0]
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'c'] = results['c'][0]
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'rsquared'] = results['r2']
            lm_corrEON.loc['TI_regression_corrTI_RSD_Ref', 'rmse'] = results['rmse'][0]

            #Compute model regression results
            results = get_modelRegression(inputdataEON, 'corrTI_RSD_TI','Ref_TI')

            TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdataEON)
            rep_TI_results = get_representative_TI_15mps(inputdataEON)
            TIbybin = get_TI_bybin(inputdataEON)
            count = get_count_per_WSbin(inputdataEON, 'RSD_WS')
            total_stats = get_description_stats(inputdataEON)


            TI_MBEList.append(TI_MBE_j_)
            TI_DiffList.append(TI_Diff_j_)
            TI_RMSEList.append(TI_RMSE_j_)
            rep_TI_resultsList.append(rep_TI_results)
            TIBinList.append(TIbybin)
            countList.append(count)
            total_StatsList.append(total_stats)
            lm_CorrList.append(lm_corrEON)
            correctionTag.append(i)

        if i == 'ZX':
            if 'corrTI_RSD_TI' in inputdata.columns:
                print("Corrected TI already present in data. ZX correction not performed.")
            else:
                inputdataZX = perform_zx_correction(inputdata.copy(), height)
                results = get_modelRegression(inputdataZX, 'corrTI_RSD_TI', 'Ref_TI')

                print("ZX correction applied:\n" + str(round(zx_correct(height), 4)) + \
                      " for station height " + str(height) + " meters")
                lm_corrZX = pd.DataFrame(columns=['sensor', 'height', 'correction', 'm', \
                                                'c', 'rsquared'])
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'sensor'] = sensor
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'height'] = height
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'correction'] = 'ZX'
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'm'] = results['m'][0]
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'c'] = results['c'][0]
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'rsquared'] = results['r2']
                lm_corrZX.loc['TI_regression_corrTI_RSD_Ref', 'rmse'] = results['rmse'][0]


                TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdataZX)
                rep_TI_results = get_representative_TI_15mps(inputdataZX)
                TIbybin = get_TI_bybin(inputdataZX)
                count = get_count_per_WSbin(inputdataZX, 'RSD_WS')
                total_stats = get_description_stats(inputdataZX)

                TI_MBEList.append(TI_MBE_j_)
                TI_DiffList.append(TI_Diff_j_)
                TI_RMSEList.append(TI_RMSE_j_)
                rep_TI_resultsList.append(rep_TI_results)
                TIBinList.append(TIbybin)
                countList.append(count)
                total_StatsList.append(total_stats)
                lm_CorrList.append(lm_corrZX)
                correctionTag.append(i)

        if i == 'VAISALA':
            if 'corrTI_RSD_TI' in inputdata.columns:
                print("Vaisala-Triton corrected TI already present in data. Reporting results.")
                lm_corrVaisala = pd.DataFrame(columns=['sensor', 'height', 'correction', 'm', \
                                                       'c', 'rsquared', 'ws_diff'])
                lm_corrVaisala.loc['TI_regression_corrTI_RSD_Ref', 'sensor'] = sensor
                lm_corrVaisala.loc['TI_regression_corrTI_RSD_Ref', 'height'] = height
                lm_corrVaisala.loc['TI_regression_corrTI_RSD_Ref', 'correction'] = 'Vaisala'

                TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdata.copy())
                rep_TI_results = get_representative_TI_15mps(inputdata.copy())
                TIbybin = get_TI_bybin(inputdata.copy())
                count = get_count_per_WSbin(inputdata.copy(), 'RSD_WS')
                total_stats = get_description_stats(inputdata.copy())

                TI_MBEList.append(TI_MBE_j_)
                TI_DiffList.append(TI_Diff_j_)
                TI_RMSEList.append(TI_RMSE_j_)
                rep_TI_resultsList.append(rep_TI_results)
                TIBinList.append(TIbybin)
                countList.append(count)
                total_StatsList.append(total_stats)
                lm_CorrList.append(lm_corrVaisala)
                correctionTag.append(i)
            else:
                print("Vaisala - Triton correction applied to input datanot applied because no data was labeled as corrected\n")
                correctionTag.append('NoCorrection')
                lm_corr = pd.DataFrame(columns=['sensor', 'height', 'correction', 'm', \
                                            'c', 'rsquared'])
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'sensor'] = None
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'height'] = None
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'correction'] = 'No correction'
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'm'] = None
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'c'] = None
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'rsquared'] = None
                lm_corr.loc['TI_regression_corrTI_RSD_Ref', 'rmse'] = None

                #Compute model regression results
                TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdata)
                rep_TI_results = get_representative_TI_15mps(inputdata)
                TIbybin = get_TI_bybin(inputdata)
                count = get_count_per_WSbin(inputdata, 'Ane2_WS')
                total_stats = get_description_stats(inputdata)

                TI_MBEList.append(TI_MBE_j_)
                TI_DiffList.append(TI_Diff_j_)
                TI_RMSEList.append(TI_RMSE_j_)
                rep_TI_resultsList.append(rep_TI_results)
                TIBinList.append(TIbybin)
                countList.append(count)
                total_StatsList.append(total_stats)
                lm_CorrList.append(lm_corr)
                
        if i == 'ZXOld':
            if 'corrTI_RSD_TI' in inputdata.columns:
                print("Corrected TI already present in data. ZX correction not performed.")
            else:
                inputdataZXOld = perform_zx_correction_oldAB(inputdata.copy(), height)
                results = get_modelRegression(inputdataZXOld, 'corrTI_RSD_TI', 'Ref_TI')

                print("Old ZX correction applied:\n" + str(round(zx_correct(height), 4)) + \
                      " for station height " + str(height) + " meters")
                lm_corrZXOld = pd.DataFrame(columns=['sensor', 'height', 'correction', 'm', \
                                                'c', 'rsquared'])
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'sensor'] = sensor
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'height'] = height
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'correction'] = 'ZX'
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'm'] = results['m'][0]
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'c'] = results['c'][0]
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'rsquared'] = results['r2']
                lm_corrZXOld.loc['TI_regression_corrTI_RSD_Ref', 'rmse'] = results['rmse'][0]


                TI_MBE_j_, TI_Diff_j_, TI_RMSE_j_ = get_TI_MBE_Diff_j(inputdataZXOld)
                rep_TI_results = get_representative_TI_15mps(inputdataZXOld)
                TIbybin = get_TI_bybin(inputdataZXOld)
                count = get_count_per_WSbin(inputdataZXOld, 'RSD_WS')
                total_stats = get_description_stats(inputdataZXOld)

                TI_MBEList.append(TI_MBE_j_)
                TI_DiffList.append(TI_Diff_j_)
                TI_RMSEList.append(TI_RMSE_j_)
                rep_TI_resultsList.append(rep_TI_results)
                TIBinList.append(TIbybin)
                countList.append(count)
                total_StatsList.append(total_stats)
                lm_CorrList.append(lm_corrZXOld)
                correctionTag.append(i)
        if i == 'L-TERRA':
            pass

    write_all_resultstofile(reg_results, allTIRegressResults, TI_MBEList,TI_DiffList, TI_RMSEList,rep_TI_resultsList,\
                            TIBinList, countList, total_StatsList, results_filename,\
                            lm_CorrList, correctionTag, siteMetadata, filterMetadata)

